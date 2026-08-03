# -*- coding: utf-8 -*-
# Part of gdrive_odoo_sync. Licensed under LGPL-3.
"""Lane B — the ``.xlsx`` reader (SPEC §4.7).

WHY a second reader exists at all
=================================
The Sheets API **cannot** read an ``.xlsx`` sitting in Drive: the file has no
``spreadsheetId``, and ``spreadsheets.get`` on its Drive file id returns 404.
Neither can the obvious workaround be used — converting it to a native Sheet
needs write scope, and the service account's 0 GB quota makes ``files.copy``
fail with ``storageQuotaExceeded`` unless the copy is parented into a shared
drive. So the supported path is download plus local parse, and this module
tests that parse.

Two openpyxl behaviours carry real risk:

* ``data_only=True`` returns the value Excel *cached* the last time it
  recalculated. If the file was written by a tool that never recalculated, every
  formula cell comes back as ``None`` — which is indistinguishable from an empty
  cell. Reading those as empty would silently blank an entire column, so they
  are flagged (``XLSX_NO_CACHED_VALUES``) and the affected cells are quarantined
  rather than read.
* ``read_only=True`` streams, which is what keeps a multi-megabyte cashflow
  workbook from being materialized twice in RAM inside a cron worker.

And one structural gap: an xlsx worksheet has **no stable gid**. The dataset
identity is therefore the negative surrogate ``-(1 + worksheet_index)``, which
is why a worksheet rename is genuinely ambiguous with delete-plus-create and
has to be resolved by title first, then index.
"""

import io
import unittest
from unittest import mock

from odoo.tests.common import BaseCase

from odoo.addons.gdrive_odoo_sync.services.errors import (
    CODE_XLSX_NO_CACHED_VALUES,
    CODE_XLSX_SCAN_INCOMPLETE,
    GDrivePermanentError,
)
from odoo.addons.gdrive_odoo_sync.services.mimetypes import is_legacy_xls, is_spreadsheet_blob
from odoo.addons.gdrive_odoo_sync.services.xlsx_reader import XlsxReader

try:
    import openpyxl
except ImportError:  # pragma: no cover - openpyxl ships with Odoo
    openpyxl = None

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
XLS_MIME = "application/vnd.ms-excel"

requires_openpyxl = unittest.skipUnless(openpyxl is not None, "openpyxl is not installed")


def _workbook(sheets):
    """Build an in-memory ``.xlsx`` from ``{title: [[cell, ...], ...]}`` pairs."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for title, rows in sheets:
        ws = wb.create_sheet(title=title)
        for row in rows:
            ws.append(list(row))
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class TestMimeGuards(BaseCase):
    """Legacy ``.xls`` is unsupported in v1 and must be refused, not attempted."""

    def test_xlsx_is_recognised_as_a_spreadsheet_blob(self):
        self.assertTrue(is_spreadsheet_blob(XLSX_MIME))

    def test_legacy_xls_is_recognised_and_is_not_a_supported_spreadsheet(self):
        self.assertTrue(is_legacy_xls(XLS_MIME))
        self.assertFalse(is_spreadsheet_blob(XLS_MIME))

    def test_reader_refuses_a_legacy_xls_payload(self):
        # openpyxl reads only OOXML; feeding it a BIFF stream produces a
        # confusing library error rather than a clear "unsupported" message.
        with self.assertRaises(GDrivePermanentError):
            XlsxReader().read(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1legacy-ole2-header")

    def test_reader_refuses_garbage(self):
        with self.assertRaises(GDrivePermanentError):
            XlsxReader().read(b"this is not a workbook")


@requires_openpyxl
class TestWorksheetEnumeration(BaseCase):
    """The reader presents the same tab shape as the Sheets reader."""

    def test_every_worksheet_is_returned_in_order(self):
        data = _workbook([("Summary", [["a"]]), ("Detail", [["b"]]), ("Notes", [["c"]])])
        tabs = XlsxReader().read(data)
        self.assertEqual([t["tab_title"] for t in tabs], ["Summary", "Detail", "Notes"])
        self.assertEqual([t["tab_index"] for t in tabs], [0, 1, 2])

    def test_shape_matches_the_sheets_reader(self):
        # The key set is the one SheetsReader._tab_meta / _tab_values emit, so
        # lane D stages a native Sheet and an uploaded workbook through one code
        # path. Asserting 'index'/'title' here asserted a shape *neither* reader
        # produces, which is worse than no test.
        data = _workbook([("Summary", [["a"]])])
        tab = XlsxReader().read(data)[0]
        for key in ("source_kind", "sheet_gid", "tab_index", "tab_title", "hidden",
                    "sheet_type", "rows", "row_count", "col_count", "used_range",
                    "read_complete", "warnings", "error_cells"):
            with self.subTest(key=key):
                self.assertIn(key, tab)

    def test_negative_surrogate_gid_formula(self):
        # xlsx worksheets have no stable gid, so the identity is a negative
        # surrogate. It must never collide with a real Google Sheets gid, which
        # is always >= 0 (and 0 is a real, common value).
        data = _workbook([("A", [["1"]]), ("B", [["2"]]), ("C", [["3"]])])
        tabs = XlsxReader().read(data)
        gids = [t["sheet_gid"] for t in tabs]
        self.assertEqual(gids, [-1, -2, -3])
        self.assertTrue(all(g < 0 for g in gids))
        self.assertEqual(len(set(gids)), len(gids))

    def test_used_range_is_the_populated_extent(self):
        data = _workbook([("S", [["h1", "h2"], ["a", "b"], ["c", "d"]])])
        tab = XlsxReader().read(data)[0]
        self.assertIn("S", tab["used_range"])
        self.assertTrue(tab["used_range"].endswith("B3"), tab["used_range"])


@requires_openpyxl
class TestRowShaping(BaseCase):
    """Rows are right-padded so every downstream index is safe."""

    def test_ragged_rows_are_padded_to_the_widest(self):
        data = _workbook([("S", [[1, 2, 3], [4], [5, 6]])])
        rows = XlsxReader().read(data)[0]["rows"]
        self.assertEqual([len(r) for r in rows], [3, 3, 3])

    def test_interior_empties_are_preserved(self):
        data = _workbook([("S", [["a", None, "c"]])])
        rows = XlsxReader().read(data)[0]["rows"]
        self.assertEqual(len(rows[0]), 3)
        self.assertIn(rows[0][1], (None, ""))

    def test_empty_worksheet_yields_no_rows(self):
        data = _workbook([("Empty", [])])
        tab = XlsxReader().read(data)[0]
        self.assertEqual(tab["rows"], [])

    def test_values_keep_their_python_types(self):
        # Numbers must arrive as numbers so NUM_CANON sees a float rather than a
        # locale-formatted string.
        data = _workbook([("S", [["name", "amount"], ["ACME", 1234.5]])])
        rows = XlsxReader().read(data)[0]["rows"]
        self.assertEqual(rows[1][0], "ACME")
        self.assertIsInstance(rows[1][1], (int, float))
        self.assertAlmostEqual(float(rows[1][1]), 1234.5)


@requires_openpyxl
class TestFormulaCaching(BaseCase):
    """``data_only=True`` returns ``None`` for never-recalculated formulas."""

    def test_uncached_formula_is_flagged_not_read_as_empty(self):
        # openpyxl writes the formula string with no cached value, which is
        # exactly the shape a file produced by a non-Excel tool has.
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Calc"
        ws["A1"] = "total"
        ws["A2"] = "=SUM(B1:B9)"
        buffer = io.BytesIO()
        wb.save(buffer)

        reader = XlsxReader()
        tab = reader.read(buffer.getvalue())[0]
        # The reader reports this through read_complete / warnings / error_cells.
        # The old assertion read `tab.get("no_cached_values")` and
        # `tab.get("flags")` — keys it never emits — so `None or ("..." in [])`
        # was False and the test failed even though detection had worked.
        self.assertFalse(tab["read_complete"],
                         "an uncached formula must disarm deletions: %r" % (tab,))
        self.assertIn(CODE_XLSX_NO_CACHED_VALUES, [w["code"] for w in tab["warnings"]])
        self.assertTrue(tab["error_cells"])
        self.assertEqual(tab["error_cells"][0]["code"], CODE_XLSX_NO_CACHED_VALUES)

    def test_plain_values_are_not_flagged(self):
        # The sibling test above used to pass only by accident (both .get() calls
        # returned None on a dict that has neither key), so it asserted nothing.
        data = _workbook([("S", [["a", 1], ["b", 2]])])
        tab = XlsxReader().read(data)[0]
        self.assertTrue(tab["read_complete"])
        self.assertEqual(tab["warnings"], [])
        self.assertEqual(tab["error_cells"], [])


@requires_openpyxl
class TestFormulaScanFailure(BaseCase):
    """A pre-scan we could not run must not present as a clean read."""

    def test_a_failed_prescan_reports_every_worksheet_incomplete(self):
        # A workbook that opens under data_only=True but trips openpyxl in the
        # data_only=False pass used to return an empty formula map, which is
        # byte-for-byte the same answer as "this workbook has no formulas": every
        # worksheet was reported read_complete=True, uncached formula cells staged
        # as '', and the delete planner stayed armed. The mechanism that disarms
        # deletions was silently disarming itself.
        data = _workbook([("S", [["a", 1]]), ("T", [["b", 2]])])
        reader = XlsxReader()
        real_load = openpyxl.load_workbook

        def flaky(stream, **kwargs):
            if not kwargs.get("data_only", False):
                raise ValueError("simulated defined-name parse failure")
            return real_load(stream, **kwargs)

        with mock.patch.object(openpyxl, "load_workbook", flaky):
            tabs = reader.read(data)

        self.assertEqual(len(tabs), 2)
        for tab in tabs:
            with self.subTest(tab=tab["tab_title"]):
                self.assertFalse(tab["read_complete"])
                self.assertIn(CODE_XLSX_SCAN_INCOMPLETE, [w["code"] for w in tab["warnings"]])


@requires_openpyxl
class TestLargeWorkbookStreaming(BaseCase):
    """``read_only=True`` is what makes a multi-megabyte workbook safe in a cron."""

    def test_a_few_thousand_rows_read_without_incident(self):
        rows = [["r%d" % i, i, i * 1.5] for i in range(5000)]
        data = _workbook([("Big", [["k", "n", "f"]] + rows)])
        tab = XlsxReader().read(data)[0]
        self.assertEqual(len(tab["rows"]), 5001)
        self.assertEqual(tab["rows"][-1][0], "r4999")

    def test_multiple_reads_of_the_same_bytes_agree(self):
        data = _workbook([("S", [["a", 1], ["b", 2]])])
        reader = XlsxReader()
        self.assertEqual(reader.read(data)[0]["rows"], reader.read(data)[0]["rows"])
